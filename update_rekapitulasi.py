import json
import re
from collections import defaultdict
from openpyxl import load_workbook

INPUT_JSON = "DUKPNSFULL.json"
WORKBOOK_FILE = "DUK BAPENDA KOTA MEDAN 2026 1 Maret.xlsx"
SHEET_NAME = "REKAPITULASI"


UNIT_ROWS = {
    "SEKRETARIAT": 4,
    "BIDANG 1": 5,
    "BIDANG 2": 6,
    "BIDANG 3": 7,
    "BIDANG 4": 8,
    "UPT 1": 9,
    "UPT 2": 10,
    "UPT 3": 11,
    "UPT 4": 12,
    "UPT 5": 13,
    "UPT 6": 14,
    "UPT 7": 15,
}

GOL_ROWS = {
    "IV/C": 31,
    "IV/B": 32,
    "IV/A": 33,
    "III/D": 34,
    "III/C": 35,
    "III/B": 36,
    "III/A": 37,
    "II/D": 38,
    "II/C": 39,
    "II/B": 40,
    "II/A": 41,
    "I/D": 42,
    "I/C": 43,
    "I/B": 44,
    "I/A": 45,
}

PEND_ROWS = {
    "S3": 51,
    "S2": 52,
    "S1": 53,
    "DIPLOMA": 54,
    "SLTA": 55,
    "SLTP": 56,
    "SD": 57,
}


def norm(value):
    return re.sub(r"\s+", " ", str(value or "").strip().upper())


def latest_item(value):
    if isinstance(value, list) and value:
        return value[-1]
    return {}


def is_male(emp):
    return norm(emp.get("jenisKelamin")).startswith("LAKI")


def classify_unit(emp):
    jab = latest_item(emp.get("riwayatJabatan"))
    skpd = norm(jab.get("skpd") or emp.get("skpd"))
    m = re.search(r"WILAYAH\s+(I|II|III|IV|V|VI|VII)\b", skpd)
    if m:
        roman_to_num = {"I": "1", "II": "2", "III": "3", "IV": "4", "V": "5", "VI": "6", "VII": "7"}
        return f"UPT {roman_to_num[m.group(1)]}"

    if "PENGEMBANGAN DAN PENGENDALIAN" in skpd:
        return "BIDANG 4"
    if any(k in skpd for k in ["PARKIR", "REKLAME", "PENERANGAN JALAN", "AIR TANAH", "WALET", "RETRIBUSI"]):
        return "BIDANG 3"
    if any(k in skpd for k in ["HOTEL", "RESTORAN", "HIBURAN"]):
        return "BIDANG 2"
    if any(k in skpd for k in ["BEA PEROLEHAN HAK ATAS TANAH", "PAJAK BUMI DAN BANGUNAN"]):
        return "BIDANG 1"
    return "SEKRETARIAT"


def normalize_gol(gol):
    g = norm(gol).replace(".", "/").replace(" ", "")
    m = re.match(r"^(IV|III|II|I)[/ ]?([A-E])$", g)
    if m:
        return f"{m.group(1)}/{m.group(2)}"
    return ""


def classify_pendidikan(emp):
    rp = latest_item(emp.get("riwayatPendidikan"))
    text = " ".join(
        [
            norm(rp.get("tingkatPendidikan")),
            norm(rp.get("detailPendidikan")),
            norm(rp.get("jurusan")),
            norm(rp.get("namaSekolah")),
        ]
    )

    if any(k in text for k in ["STRATA - 3", "STRATA 3", "S3", "DOKTOR"]):
        return "S3"
    if any(k in text for k in ["STRATA - 2", "STRATA 2", "S2", "MAGISTER"]):
        return "S2"
    if any(k in text for k in ["STRATA - 1", "STRATA 1", "S1", "SARJANA"]) and "SARJANA TERAPAN" not in text:
        return "S1"
    if any(k in text for k in ["DIPLOMA", "D-I", "D-II", "D-III", "D-IV", "D1", "D2", "D3", "D4", "SARJANA TERAPAN", "AHLI MADYA"]):
        return "DIPLOMA"
    if any(k in text for k in ["SLTA", "SMA", "SMK", "STM", "AMK", "ALIYAH", "MA "]):
        return "SLTA"
    if any(k in text for k in ["SLTP", "SMP", "MTS"]):
        return "SLTP"
    if " SD" in f" {text}" or text.startswith("SD "):
        return "SD"
    return ""


def main():
    with open(INPUT_JSON, "r") as f:
        data = json.load(f)

    unit_counts = {k: {"pns": 0, "pppk_full": 0, "pppk_part": 0} for k in UNIT_ROWS}
    jabatan_counts = {
        "ESELON II": {"total": 0, "male": 0},
        "ESELON III": {"total": 0, "male": 0},
        "ESELON IV": {"total": 0, "male": 0},
        "FUNGSIONAL": {"total": 0, "male": 0},
        "STAF": {"total": 0, "male": 0},
    }
    gol_counts = {k: {"total": 0, "male": 0} for k in GOL_ROWS}
    pend_pns = {k: {"total": 0, "male": 0} for k in PEND_ROWS}
    pend_pppk = {k: {"total": 0, "male": 0} for k in PEND_ROWS}

    for emp in data:
        status = norm(emp.get("statusKepegawaian"))
        unit = classify_unit(emp)
        male = is_male(emp)

        if status in {"PNS", "CPNS"}:
            unit_counts[unit]["pns"] += 1

            jab = latest_item(emp.get("riwayatJabatan"))
            eselon_raw = norm(jab.get("eselon"))
            m_es = re.match(r"^(I|II|III|IV|V)\b", eselon_raw.replace(".", " "))
            eselon = m_es.group(1) if m_es else ""
            tipe = norm(jab.get("tipeJabatan"))

            if eselon == "II":
                key = "ESELON II"
            elif eselon == "III":
                key = "ESELON III"
            elif eselon == "IV":
                key = "ESELON IV"
            elif "FUNGSIONAL" in tipe and "UMUM" not in tipe and "PELAKSANA" not in tipe:
                key = "FUNGSIONAL"
            else:
                key = "STAF"
            jabatan_counts[key]["total"] += 1
            jabatan_counts[key]["male"] += 1 if male else 0

            gol = normalize_gol(emp.get("golonganTerakhir"))
            if gol in gol_counts:
                gol_counts[gol]["total"] += 1
                gol_counts[gol]["male"] += 1 if male else 0

            pend = classify_pendidikan(emp)
            if pend in pend_pns:
                pend_pns[pend]["total"] += 1
                pend_pns[pend]["male"] += 1 if male else 0

        elif status == "PPPK":
            # Data JSON tidak memuat penanda penuh/paruh waktu.
            # Semua PPPK direkap sebagai paruh waktu agar tetap terhitung di ASN.
            unit_counts[unit]["pppk_part"] += 1

            pend = classify_pendidikan(emp)
            if pend in pend_pppk:
                pend_pppk[pend]["total"] += 1
                pend_pppk[pend]["male"] += 1 if male else 0

    wb = load_workbook(WORKBOOK_FILE)
    ws = wb[SHEET_NAME]

    for unit, row in UNIT_ROWS.items():
        ws.cell(row=row, column=3).value = unit_counts[unit]["pns"]
        ws.cell(row=row, column=4).value = unit_counts[unit]["pppk_full"]
        ws.cell(row=row, column=5).value = unit_counts[unit]["pppk_part"]

    jabatan_rows = {
        "ESELON II": 21,
        "ESELON III": 22,
        "ESELON IV": 23,
        "FUNGSIONAL": 24,
        "STAF": 25,
    }
    for key, row in jabatan_rows.items():
        total = jabatan_counts[key]["total"]
        male = jabatan_counts[key]["male"]
        ws.cell(row=row, column=3).value = total
        ws.cell(row=row, column=5).value = male
        ws.cell(row=row, column=6).value = total - male

    for gol, row in GOL_ROWS.items():
        total = gol_counts[gol]["total"]
        male = gol_counts[gol]["male"]
        ws.cell(row=row, column=3).value = total
        ws.cell(row=row, column=4).value = male
        ws.cell(row=row, column=5).value = total - male

    for row in range(31, 46):
        ws.cell(row=row, column=10).value = 0
        ws.cell(row=row, column=11).value = 0

    for pend, row in PEND_ROWS.items():
        total = pend_pns[pend]["total"]
        male = pend_pns[pend]["male"]
        ws.cell(row=row, column=3).value = total
        ws.cell(row=row, column=4).value = male
        ws.cell(row=row, column=5).value = total - male

    for pend, row in PEND_ROWS.items():
        total = pend_pppk[pend]["total"]
        male = pend_pppk[pend]["male"]
        ws.cell(row=row, column=10).value = male
        ws.cell(row=row, column=11).value = total - male

    wb.save(WORKBOOK_FILE)

    total_pns = sum(v["pns"] for v in unit_counts.values())
    total_pppk = sum(v["pppk_full"] + v["pppk_part"] for v in unit_counts.values())
    print(f"Rekap updated: PNS/CPNS={total_pns}, PPPK={total_pppk}, total={total_pns + total_pppk}")


if __name__ == "__main__":
    main()
