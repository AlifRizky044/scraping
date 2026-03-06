import json
import re
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook

TEMPLATE_FILE = "DUK FORMAT.xlsx"
INPUT_JSON = "DUKPNSFULL.json"
OUTPUT_FILE = "DUK_PNS_CPNS_JABATAN_TERAKHIR.xlsx"
REMOVAL_LOG_FILE = "DUK_REMOVED_LOG.txt"


def normalize_text(value):
    if value is None:
        return ""
    text = str(value).strip().upper()
    text = text.replace(".", "").replace(",", "")
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_name(value):
    text = normalize_text(value)
    text = re.sub(r"[^A-Z0-9 ]", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_nip(value):
    return re.sub(r"\D", "", str(value or ""))


def golongan_rank(golongan):
    ranks = {
        "IV/E": 17,
        "IV/D": 16,
        "IV/C": 15,
        "IV/B": 14,
        "IV/A": 13,
        "III/D": 12,
        "III/C": 11,
        "III/B": 10,
        "III/A": 9,
        "II/D": 8,
        "II/C": 7,
        "II/B": 6,
        "II/A": 5,
        "I/D": 4,
        "I/C": 3,
        "I/B": 2,
        "I/A": 1,
    }
    return ranks.get(normalize_text(golongan), 0)


def extract_golongan_from_pangkat(pangkat):
    text = str(pangkat or "")
    if "(" in text and ")" in text:
        return text.split("(", 1)[1].split(")", 1)[0].strip().upper()
    return ""


def parse_date(date_str):
    text = str(date_str or "").strip()
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%Y"):
        try:
            dt = datetime.strptime(text[:10], fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    return text


def compute_bup(tanggal_lahir, eselon):
    birth = str(tanggal_lahir or "").strip()
    if not birth:
        return ""
    try:
        dt = datetime.strptime(birth[:10], "%Y-%m-%d")
    except ValueError:
        return ""

    eselon_norm = normalize_text(eselon)
    bup_age = 60 if eselon_norm.startswith("II") else 58
    bup_date = dt + relativedelta(years=bup_age, months=1)
    bup_date = bup_date.replace(day=1)
    return bup_date.strftime("%d-%m-%Y")


def safe_str(value):
    if value is None:
        return ""
    return str(value)


def get_latest_jabatan(emp):
    riwayat = emp.get("riwayatJabatan") or []
    if isinstance(riwayat, list) and riwayat:
        return riwayat[-1]
    return {}


def get_latest_pangkat(emp):
    riwayat = emp.get("riwayatPangkat") or []
    if isinstance(riwayat, list) and riwayat:
        return riwayat[-1]
    return {}


def get_first_pangkat(emp):
    riwayat = emp.get("riwayatPangkat") or []
    if isinstance(riwayat, list) and riwayat:
        return riwayat[0]
    return {}


def get_latest_pendidikan(emp):
    riwayat = emp.get("riwayatPendidikan") or []
    if isinstance(riwayat, list) and riwayat:
        return riwayat[-1]
    return {}


def get_latest_diklat(emp):
    riwayat = emp.get("riwayatDiklatStruktural") or []
    if isinstance(riwayat, list) and riwayat:
        return riwayat[-1]
    return {}


def get_emp_jabatan_norm(emp):
    return normalize_text(emp.get("jabatanTerakhir") or get_latest_jabatan(emp).get("jabatan"))


def get_upt_section(text):
    t = normalize_text(text)
    m = re.search(r"WILAYAH\s+(III|VII|VI|IV|II|I|V|1|2|3|4|5|6|7)\b", t)
    if not m:
        m = re.search(r"\bUPT\s+(III|VII|VI|IV|II|I|V|1|2|3|4|5|6|7)\b", t)
    if not m:
        return ""
    token = m.group(1)
    arabic_to_roman = {
        "1": "I",
        "2": "II",
        "3": "III",
        "4": "IV",
        "5": "V",
        "6": "VI",
        "7": "VII",
    }
    return arabic_to_roman.get(token, token)


def get_emp_upt_section(emp):
    latest = get_latest_jabatan(emp)
    skpd = latest.get("skpd") or emp.get("skpd") or ""
    return get_upt_section(skpd)


def duk_rank(emp):
    latest_pangkat = get_latest_pangkat(emp)
    latest_jabatan = get_latest_jabatan(emp)

    gol = extract_golongan_from_pangkat(latest_pangkat.get("pangkat"))
    grank = golongan_rank(gol)

    eselon = normalize_text(latest_jabatan.get("eselon"))
    eselon_rank_map = {
        "II A": 10,
        "II B": 9,
        "III A": 8,
        "III B": 7,
        "IV A": 6,
        "IV B": 5,
        "V A": 4,
        "V B": 3,
    }
    erank = eselon_rank_map.get(eselon, 0)

    try:
        mkg_tahun = int(latest_pangkat.get("mkgTahun") or 0)
    except (TypeError, ValueError):
        mkg_tahun = 0

    nama = normalize_name(emp.get("namaWithGelar") or emp.get("nama"))
    return (-erank, -grank, -mkg_tahun, nama)


def find_row_by_jabatan(ws, jabatan_pattern_norm, start_row):
    for r in range(start_row, ws.max_row + 1):
        jab = normalize_text(ws.cell(row=r, column=15).value)
        if jabatan_pattern_norm in jab:
            return r
    return None


def find_emp_by_exact_jabatan(employees, jabatan_norm):
    for emp in employees:
        if get_emp_jabatan_norm(emp) == jabatan_norm:
            return emp
    return None


def find_upt_section_append_row(ws, target_upt, start_row):
    # Find the row marker "UPT <target>" in column 1, then append at end of that block.
    start_marker_row = None
    for r in range(start_row, ws.max_row + 1):
        marker_upt = get_upt_section(safe_str(ws.cell(row=r, column=1).value))
        if marker_upt == target_upt:
            start_marker_row = r
            break
    if start_marker_row is None:
        return None

    # End is right before next UPT marker.
    next_marker_row = None
    for r in range(start_marker_row + 1, ws.max_row + 1):
        marker_upt = get_upt_section(safe_str(ws.cell(row=r, column=1).value))
        if marker_upt:
            next_marker_row = r
            break

    if next_marker_row is None:
        end_row = ws.max_row
    else:
        end_row = next_marker_row - 1

    # Insert after the last non-empty row within the UPT block.
    last_used = start_marker_row
    for r in range(start_marker_row, end_row + 1):
        if any(safe_str(ws.cell(row=r, column=c).value).strip() for c in (1, 2, 3, 15, 17)):
            last_used = r
    return last_used + 1


def ensure_kabid_hotel_present(ws, employees, used_nips, start_row):
    kabid_hotel_norm = normalize_text(
        "KEPALA BIDANG HOTEL, RESTORAN, DAN HIBURAN BADAN PENDAPATAN DAERAH KOTA MEDAN"
    )
    hotel_subbid_norm = normalize_text(
        "KEPALA SUB BIDANG TEKNIS HOTEL, RESTORAN, DAN HIBURAN BADAN PENDAPATAN DAERAH KOTA MEDAN"
    )

    # If already present, nothing to do.
    existing_row = find_row_by_jabatan(ws, kabid_hotel_norm, start_row)
    if existing_row is not None:
        return None

    emp = find_emp_by_exact_jabatan(employees, kabid_hotel_norm)
    if not emp:
        return "Kepala Bidang Hotel tidak ditemukan di JSON, tidak bisa disisipkan."

    nip = normalize_nip(emp.get("nip"))
    if nip and nip in used_nips:
        return None

    insert_before = find_row_by_jabatan(ws, hotel_subbid_norm, start_row)
    if insert_before is None:
        insert_before = start_row

    ws.insert_rows(insert_before, 1)
    values = build_row(emp)
    for c in range(1, 45):
        ws.cell(row=insert_before, column=c).value = values[c - 1]

    if nip:
        used_nips.add(nip)

    return f"Inserted Kepala Bidang Hotel at row {insert_before}."


def move_person_to_unit_block(ws, person_name_keyword, unit_keyword, jabatan_keyword, start_row):
    person_norm = normalize_text(person_name_keyword)
    unit_norm = normalize_text(unit_keyword)
    jab_norm = normalize_text(jabatan_keyword)

    src_row = None
    for r in range(start_row, ws.max_row + 1):
        nm = normalize_text(ws.cell(row=r, column=2).value)
        if person_norm in nm:
            src_row = r
            break
    if src_row is None:
        return None

    target_row = None
    target_unit_value = None
    for r in range(start_row, ws.max_row + 1):
        unit = normalize_text(ws.cell(row=r, column=17).value)
        jab = normalize_text(ws.cell(row=r, column=15).value)
        if unit_norm in unit and jab == jab_norm:
            target_row = r
            target_unit_value = ws.cell(row=r, column=17).value
            break
    if target_row is None or target_row == src_row:
        return None

    # Insert as a new row in the target block (do not replace existing occupant).
    insert_row = target_row + 1
    while insert_row <= ws.max_row:
        u = normalize_text(ws.cell(row=insert_row, column=17).value)
        j = normalize_text(ws.cell(row=insert_row, column=15).value)
        if unit_norm in u and j == jab_norm:
            insert_row += 1
            continue
        break

    src_vals = [ws.cell(row=src_row, column=c).value for c in range(1, 45)]
    ws.insert_rows(insert_row, 1)
    for c in range(1, 45):
        ws.cell(row=insert_row, column=c).value = src_vals[c - 1]
    ws.cell(row=insert_row, column=15).value = jabatan_keyword
    ws.cell(row=insert_row, column=17).value = target_unit_value

    # Clear original row so person is relocated, not duplicated.
    for c in range(1, 45):
        ws.cell(row=src_row, column=c).value = None

    return f"Inserted '{person_name_keyword}' as new row {insert_row} in target block."


def build_row(emp):
    row = [""] * 44

    row[1] = safe_str(emp.get("namaWithGelar") or emp.get("nama") or "").upper()
    nip = safe_str(emp.get("nip") or "").strip()
    row[2] = f"'{nip}" if nip else ""

    latest_pangkat = get_latest_pangkat(emp)
    first_pangkat = get_first_pangkat(emp)

    row[3] = safe_str(latest_pangkat.get("mkgTahun"))
    row[4] = safe_str(latest_pangkat.get("mkgBulan"))

    cpns_pangkat = safe_str(first_pangkat.get("pangkat"))
    row[5] = cpns_pangkat
    row[6] = extract_golongan_from_pangkat(cpns_pangkat)
    row[7] = parse_date(first_pangkat.get("tmt"))

    last_pangkat = safe_str(latest_pangkat.get("pangkat"))
    row[8] = last_pangkat
    row[9] = extract_golongan_from_pangkat(last_pangkat)
    row[10] = safe_str(latest_pangkat.get("noSk"))
    row[11] = safe_str(latest_pangkat.get("pejabatPenandatangan"))
    row[12] = parse_date(latest_pangkat.get("tmt"))

    latest_jabatan = get_latest_jabatan(emp)
    row[13] = safe_str(latest_jabatan.get("tipeJabatan"))
    row[14] = safe_str(emp.get("jabatanTerakhir") or latest_jabatan.get("jabatan") or "").upper()
    row[15] = safe_str(latest_jabatan.get("eselon"))
    row[16] = safe_str(latest_jabatan.get("skpd") or emp.get("skpd") or "").upper()
    row[17] = parse_date(latest_jabatan.get("tmt"))
    row[18] = safe_str(latest_jabatan.get("tmtSk"))
    row[19] = safe_str(latest_jabatan.get("tglSk"))
    row[20] = safe_str(latest_jabatan.get("pejabatPenetapan"))

    latest_pendidikan = get_latest_pendidikan(emp)
    row[21] = safe_str(latest_pendidikan.get("tingkatPendidikan"))
    row[22] = safe_str(latest_pendidikan.get("jurusan"))
    row[23] = safe_str(latest_pendidikan.get("namaSekolah"))
    row[24] = safe_str(latest_pendidikan.get("tanggalIjazah"))

    latest_diklat = get_latest_diklat(emp)
    row[25] = safe_str(latest_diklat.get("namaDiklat"))

    row[26] = compute_bup(emp.get("tanggalLahir"), row[15])

    row[28] = safe_str(emp.get("helper"))
    row[29] = safe_str(emp.get("statusKepegawaian") or emp.get("statuskepegawaian"))
    row[30] = safe_str(emp.get("agama"))
    row[31] = safe_str(emp.get("statusPernikahan"))
    row[32] = safe_str(emp.get("jenisKelamin"))
    row[33] = safe_str(emp.get("golDarah"))
    row[34] = safe_str(emp.get("noHp"))
    row[35] = safe_str(emp.get("email"))

    ttl = safe_str(emp.get("tempatTglLahir"))
    row[36] = ttl.replace("\n", " ")
    row[37] = safe_str(emp.get("alamat"))
    row[38] = safe_str(emp.get("provinsi"))
    row[39] = safe_str(emp.get("kota"))
    row[40] = safe_str(emp.get("kecamatan"))
    row[41] = safe_str(emp.get("kelurahan"))
    row[42] = safe_str(emp.get("rt"))
    row[43] = safe_str(emp.get("rw"))

    return row


def find_emp_for_template_row(template_nip, template_name, template_jabatan_norm, nip_map, name_map):
    if template_nip and template_nip in nip_map:
        return nip_map[template_nip]

    if not template_name:
        return None

    candidates = name_map.get(template_name, [])
    if not candidates:
        return None

    if len(candidates) == 1:
        return candidates[0]

    # Disambiguate by matching jabatan when duplicate names exist.
    for emp in candidates:
        emp_jab = get_emp_jabatan_norm(emp)
        if emp_jab == template_jabatan_norm:
            return emp

    return candidates[0]


def main():
    with open(INPUT_JSON, "r", encoding="utf-8") as f:
        raw_data = json.load(f)

    employees = []
    for emp in raw_data:
        status = normalize_text(emp.get("statusKepegawaian") or emp.get("statuskepegawaian"))
        if status in {"PNS", "CPNS"}:
            employees.append(emp)

    nip_map = {}
    name_map = {}
    for emp in employees:
        nip = normalize_nip(emp.get("nip"))
        if nip:
            nip_map[nip] = emp
        name = normalize_name(emp.get("namaWithGelar") or emp.get("nama"))
        if name:
            name_map.setdefault(name, []).append(emp)

    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active

    start_row = 7
    max_row = ws.max_row

    # If a person is removed from a slot, clear the whole row content.
    personal_cols_to_clear = list(range(1, 45))

    removed_logs = []
    moved_logs = []
    insert_logs = []
    appended_logs = []
    kept_count = 0
    used_nips = set()

    pools_by_jabatan = {}
    for emp in employees:
        jab_norm = get_emp_jabatan_norm(emp)
        if not jab_norm:
            continue
        pools_by_jabatan.setdefault(jab_norm, []).append(emp)
    for jab in pools_by_jabatan:
        pools_by_jabatan[jab].sort(key=duk_rank)

    # Capture UPT section context from template heading rows (e.g. "UPT 3").
    row_upt_section = {}
    current_upt = ""
    for r in range(start_row, max_row + 1):
        marker = safe_str(ws.cell(row=r, column=1).value)
        marker_upt = get_upt_section(marker)
        if marker_upt:
            current_upt = marker_upt
        row_upt_section[r] = current_upt

    # Protect incumbents that still have valid same-jabatan slot in template,
    # so they are not consumed by earlier generic slots with same jabatan.
    protected_nips = set()
    protected_row_by_nip = {}
    for r in range(start_row, max_row + 1):
        template_name_raw = ws.cell(row=r, column=2).value
        template_nip_raw = ws.cell(row=r, column=3).value
        template_jabatan_raw = ws.cell(row=r, column=15).value

        template_name = normalize_name(template_name_raw)
        template_nip = normalize_nip(template_nip_raw)
        template_jabatan_norm = normalize_text(template_jabatan_raw)
        if not (template_name or template_nip):
            continue

        incumbent = find_emp_for_template_row(template_nip, template_name, template_jabatan_norm, nip_map, name_map)
        if not incumbent:
            continue

        inc_nip = normalize_nip(incumbent.get("nip"))
        inc_jab = get_emp_jabatan_norm(incumbent)
        row_upt = row_upt_section.get(r, "")
        inc_upt = get_emp_upt_section(incumbent)
        compatible_upt = (not row_upt) or (not inc_upt) or (inc_upt == row_upt)
        if inc_nip and template_jabatan_norm and inc_jab == template_jabatan_norm and compatible_upt:
            protected_nips.add(inc_nip)
            protected_row_by_nip[inc_nip] = r

    for r in range(start_row, max_row + 1):
        template_name_raw = ws.cell(row=r, column=2).value
        template_nip_raw = ws.cell(row=r, column=3).value
        template_jabatan_raw = ws.cell(row=r, column=15).value
        template_unit_raw = ws.cell(row=r, column=17).value

        template_name = normalize_name(template_name_raw)
        template_nip = normalize_nip(template_nip_raw)
        template_jabatan_norm = normalize_text(template_jabatan_raw)
        template_unit_norm = normalize_text(template_unit_raw)
        row_upt = row_upt_section.get(r, "")
        row_is_upt_specific_unit = (
            "UPT PAJAK DAN RETRIBUSI WILAYAH" in template_unit_norm
            or "SUB BAGIAN TATA USAHA UPT PAJAK DAN RETRIBUSI WILAYAH" in template_unit_norm
        )

        if not (template_name or template_nip):
            continue

        incumbent = find_emp_for_template_row(template_nip, template_name, template_jabatan_norm, nip_map, name_map)
        assigned = None

        # 1) For UPT-structural rows, prioritize same-UPT candidate for this jabatan.
        if assigned is None and template_jabatan_norm and row_upt and row_is_upt_specific_unit:
            pool = pools_by_jabatan.get(template_jabatan_norm, [])
            for cand in pool:
                cand_nip = normalize_nip(cand.get("nip"))
                cand_upt = get_emp_upt_section(cand)
                if not cand_nip or cand_nip in used_nips:
                    continue
                if cand_upt != row_upt:
                    continue
                if cand_nip in protected_nips and protected_row_by_nip.get(cand_nip) != r:
                    continue
                assigned = cand
                break

        # 2) Keep incumbent in same slot if still valid and not used.
        if incumbent and assigned is None:
            inc_nip = normalize_nip(incumbent.get("nip"))
            inc_jab = get_emp_jabatan_norm(incumbent)
            inc_upt = get_emp_upt_section(incumbent)
            incumbent_upt_ok = (not row_upt) or (not inc_upt) or (inc_upt == row_upt)
            if (
                inc_nip
                and inc_nip not in used_nips
                and incumbent_upt_ok
                and (
                    (template_jabatan_norm and inc_jab == template_jabatan_norm)
                    or (not template_jabatan_norm and inc_jab)
                )
            ):
                assigned = incumbent

        # 3) If still not assigned, fill by slot jabatan pool with compatibility checks.
        if assigned is None and template_jabatan_norm:
            pool = pools_by_jabatan.get(template_jabatan_norm, [])
            for cand in pool:
                cand_nip = normalize_nip(cand.get("nip"))
                if not cand_nip or cand_nip in used_nips:
                    continue
                cand_upt = get_emp_upt_section(cand)
                # If employee has specific UPT in SKPD, keep them in same UPT section.
                if row_upt and cand_upt and cand_upt != row_upt:
                    continue
                # Do not consume UPT-specific employee for non-UPT row.
                if not row_upt and cand_upt:
                    continue
                # Do not steal protected incumbent from its own valid slot.
                if cand_nip in protected_nips and protected_row_by_nip.get(cand_nip) != r:
                    continue
                assigned = cand
                break

        # 4) For slots without jabatan in template, fallback to incumbent if exists.
        if assigned is None and not template_jabatan_norm and incumbent:
            inc_nip = normalize_nip(incumbent.get("nip"))
            inc_upt = get_emp_upt_section(incumbent)
            incumbent_upt_ok = (not row_upt) or (not inc_upt) or (inc_upt == row_upt)
            if inc_nip and inc_nip not in used_nips and incumbent_upt_ok:
                assigned = incumbent

        if assigned is None:
            old_name = safe_str(template_name_raw)
            old_nip = safe_str(template_nip_raw)
            old_jabatan = safe_str(template_jabatan_raw)
            reason = "Tidak ada pegawai aktif (PNS/CPNS) yang bisa dipasang pada slot jabatan ini."
            if incumbent:
                inc_nip = normalize_nip(incumbent.get("nip"))
                inc_jab_raw = incumbent.get("jabatanTerakhir") or get_latest_jabatan(incumbent).get("jabatan") or ""
                if inc_nip and inc_nip in used_nips:
                    moved_logs.append(
                        f"Row {r}: '{old_name}' dipindahkan ke slot lain sesuai jabatan terbaru, "
                        "slot lama dikosongkan."
                    )
                    for c in personal_cols_to_clear:
                        ws.cell(row=r, column=c).value = None
                    continue
                if not inc_jab_raw:
                    reason = "Pegawai lama ada, tetapi jabatanTerakhir kosong."
                elif template_jabatan_norm and get_emp_jabatan_norm(incumbent) != template_jabatan_norm:
                    reason = (
                        "Pegawai lama pindah jabatan dan tidak ada pengganti pada slot ini. "
                        f"Jabatan terbaru='{inc_jab_raw}'."
                    )
            removed_logs.append(f"Row {r}: {old_name} | {old_nip} | {old_jabatan} -> {reason}")

            for c in personal_cols_to_clear:
                ws.cell(row=r, column=c).value = None
            continue

        assigned_nip = normalize_nip(assigned.get("nip"))
        if assigned_nip:
            used_nips.add(assigned_nip)

        values = build_row(assigned)
        # Keep structural slot from template so hierarchy/unit placement stays intact.
        if template_jabatan_raw:
            values[14] = safe_str(template_jabatan_raw)
        if template_unit_raw:
            values[16] = safe_str(template_unit_raw)

        for c in range(1, 45):
            ws.cell(row=r, column=c).value = values[c - 1]

        if incumbent and assigned_nip and normalize_nip(incumbent.get("nip")) != assigned_nip:
            moved_logs.append(
                f"Row {r}: slot '{safe_str(template_jabatan_raw)}' diganti -> "
                f"'{safe_str(assigned.get('namaWithGelar') or assigned.get('nama'))}'"
            )

        kept_count += 1

    insert_msg = ensure_kabid_hotel_present(ws, employees, used_nips, start_row)
    if insert_msg:
        insert_logs.append(insert_msg)

    # Append remaining JSON employees who are still unplaced so nobody is lost.
    remaining = []
    for emp in employees:
        nip = normalize_nip(emp.get("nip"))
        if nip and nip not in used_nips:
            remaining.append(emp)
    remaining.sort(key=lambda e: (get_emp_upt_section(e), get_emp_jabatan_norm(e), duk_rank(e)))

    last_data_row = start_row - 1
    for r in range(start_row, ws.max_row + 1):
        if any(safe_str(ws.cell(row=r, column=c).value).strip() for c in (1, 2, 3, 15, 17)):
            last_data_row = r

    for emp in remaining:
        emp_upt = get_emp_upt_section(emp)
        if emp_upt:
            append_row = find_upt_section_append_row(ws, emp_upt, start_row)
            if append_row is None:
                append_row = last_data_row + 1
        else:
            append_row = last_data_row + 1

        ws.insert_rows(append_row, 1)
        values = build_row(emp)
        for c in range(1, 45):
            ws.cell(row=append_row, column=c).value = values[c - 1]
        if emp_upt:
            appended_logs.append(
                f"Row {append_row}: insert UPT {emp_upt} -> '{safe_str(emp.get('namaWithGelar') or emp.get('nama'))}'"
            )
        else:
            appended_logs.append(
                f"Row {append_row}: append '{safe_str(emp.get('namaWithGelar') or emp.get('nama'))}'"
            )
        nip = normalize_nip(emp.get("nip"))
        if nip:
            used_nips.add(nip)

        if append_row > last_data_row:
            last_data_row = append_row

    # Explicit user placement override.
    dimpos_move = move_person_to_unit_block(
        ws,
        person_name_keyword="DIMPOS YUSTINUS SORMIN",
        unit_keyword="BIDANG PENGEMBANGAN DAN PENGENDALIAN PAJAK DAN RETRIBUSI DAERAH",
        jabatan_keyword="PENELAAH TEKNIS KEBIJAKAN",
        start_row=start_row,
    )
    if dimpos_move:
        moved_logs.append(dimpos_move)

    # Renumber only existing people rows; removed rows get blank number.
    nomor = 1
    for r in range(start_row, ws.max_row + 1):
        name = safe_str(ws.cell(row=r, column=2).value).strip()
        if name:
            ws.cell(row=r, column=1).value = nomor
            nomor += 1
        else:
            ws.cell(row=r, column=1).value = None

    wb.save(OUTPUT_FILE)

    with open(REMOVAL_LOG_FILE, "w", encoding="utf-8") as f:
        f.write("Log pengisian ulang berdasarkan jabatanTerakhir + urutan template\n")
        f.write(f"Total slot terisi: {kept_count}\n")
        f.write(f"Total slot kosong (dihapus): {len(removed_logs)}\n")
        f.write(f"Total slot diganti personil: {len(moved_logs)}\n\n")
        if appended_logs:
            f.write(f"Total ditambahkan di bawah: {len(appended_logs)}\n")
            f.write("=== TAMBAHAN DI BAWAH TEMPLATE ===\n")
            for line in appended_logs:
                f.write(line + "\n")
            f.write("\n")
        if insert_logs:
            f.write("=== SLOT DISISIPKAN ===\n")
            for line in insert_logs:
                f.write(line + "\n")
            f.write("\n")
        if moved_logs:
            f.write("=== SLOT DIGANTI PERSONIL ===\n")
            for line in moved_logs:
                f.write(line + "\n")
            f.write("\n")
        f.write("=== SLOT KOSONG / DIHAPUS ===\n")
        f.write(f"Total dihapus: {len(removed_logs)}\n\n")
        for line in removed_logs:
            f.write(line + "\n")

    print(f"Generated {OUTPUT_FILE}")
    print(f"Kept rows: {kept_count}")
    print(f"Removed rows: {len(removed_logs)}")
    print(f"Replaced slots: {len(moved_logs)}")
    print(f"Inserted slots: {len(insert_logs)}")
    print(f"Appended rows: {len(appended_logs)}")
    print(f"Removal log: {REMOVAL_LOG_FILE}")


if __name__ == "__main__":
    main()
