import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path('/Users/nevv/Documents/scraping')
INPUT = BASE / 'DUKPNSFULLmaret.json'
OUTPUT = BASE / 'Rekap_Jumlah_ASN_BAPENDA_Menurut_Jabatan_Maret.xlsx'

with INPUT.open('r', encoding='utf-8') as f:
    data = json.load(f)

excluded_names = {
    'ANNA SARI HASIBUAN',
}

rows = [
    r for r in data
    if r.get('statusKepegawaian') in {'PNS', 'CPNS'}
    and (r.get('nama') or '').strip().upper() not in excluded_names
]

fungsional_titles = {
    'ANALIS KEUANGAN PUSAT DAN DAERAH MUDA',
    'PERENCANA MUDA',
    'ARSIPARIS PERTAMA',
    'ANALIS SUMBER DAYA MANUSIA APARATUR MUDA',
}

categories = [
    {
        'no': 1,
        'eselon': 'Eselon II',
        'keterangan': 'Kepala Badan',
        'match': lambda j: j.startswith('KEPALA BADAN '),
    },
    {
        'no': 2,
        'eselon': 'Eselon III',
        'keterangan': 'Sekretaris ( 1 Org ), Kepala Bidang ( 4 Org )',
        'match': lambda j: j.startswith('SEKRETARIS BADAN ') or j.startswith('KEPALA BIDANG '),
    },
    {
        'no': 3,
        'eselon': 'Eselon IV',
        'keterangan': 'Kasubag ( 1 Org ), Kasubbid ( 8 Org ), Ka.UPT ( 7 Org ), Kasubbag TU UPT ( 7 Org)',
        'match': lambda j: (
            j.startswith('KEPALA SUB BAGIAN UMUM ')
            or j.startswith('KEPALA SUB BIDANG ')
            or j.startswith('KEPALA UPT ')
            or j.startswith('KEPALA SUB BAGIAN TATA USAHA UPT ')
        ),
    },
    {
        'no': 4,
        'eselon': 'Fungsional',
        'keterangan': 'ANALIS KEUANGAN PUSAT DAN DAERAH MUDA ( 1 Org), PERENCANA MUDA ( 1 Org), ARSIPARIS PERTAMA ( 1 Org), ANALIS SUMBER DAYA MANUSIA APARATUR MUDA ( 1 Org)',
        'match': lambda j: j in fungsional_titles,
    },
]

assigned_indices = set()
summary = []
detail_rows = []

for category in categories:
    members = []
    for idx, row in enumerate(rows):
        if idx in assigned_indices:
            continue
        jabatan = (row.get('jabatanTerakhir') or '').strip().upper()
        if category['match'](jabatan):
            members.append(row)
            assigned_indices.add(idx)

    laki = sum(1 for r in members if r.get('jenisKelamin') == 'Laki-laki')
    perempuan = sum(1 for r in members if r.get('jenisKelamin') == 'Perempuan')
    summary.append({
        'no': category['no'],
        'eselon': category['eselon'],
        'jumlah': len(members),
        'keterangan': category['keterangan'],
        'laki_laki': laki,
        'perempuan': perempuan,
    })
    for member in members:
        detail_rows.append({
            'kategori': category['eselon'],
            'nama': member.get('nama', ''),
            'jenis_kelamin': member.get('jenisKelamin', ''),
            'status_kepegawaian': member.get('statusKepegawaian', ''),
            'jabatan_terakhir': member.get('jabatanTerakhir', ''),
            'skpd': member.get('skpd', ''),
        })

staff_members = [row for idx, row in enumerate(rows) if idx not in assigned_indices]
summary.append({
    'no': 5,
    'eselon': 'Staf',
    'jumlah': len(staff_members),
    'keterangan': 'Meliputi Staf yang ada di Kantor BAPENDA dan UPT',
    'laki_laki': sum(1 for r in staff_members if r.get('jenisKelamin') == 'Laki-laki'),
    'perempuan': sum(1 for r in staff_members if r.get('jenisKelamin') == 'Perempuan'),
})
for member in staff_members:
    detail_rows.append({
        'kategori': 'Staf',
        'nama': member.get('nama', ''),
        'jenis_kelamin': member.get('jenisKelamin', ''),
        'status_kepegawaian': member.get('statusKepegawaian', ''),
        'jabatan_terakhir': member.get('jabatanTerakhir', ''),
        'skpd': member.get('skpd', ''),
    })

total_jumlah = sum(row['jumlah'] for row in summary)
total_laki = sum(row['laki_laki'] for row in summary)
total_perempuan = sum(row['perempuan'] for row in summary)

wb = Workbook()
ws = wb.active
ws.title = 'Rekap Jabatan'

thin = Side(style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill('solid', fgColor='D9EAF7')
title_fill = PatternFill('solid', fgColor='B7DEE8')
bold = Font(bold=True)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

ws.merge_cells('A1:F1')
ws['A1'] = 'Tabel 1.2. Jumlah ASN Badan Pendapatan Daerah Kota Medan Menurut Jabatan'
ws['A1'].font = Font(bold=True, size=14)
ws['A1'].alignment = center
ws['A1'].fill = title_fill

headers = ['NO', 'ESELON', 'JUMLAH', 'KETERANGAN', 'LAKI-LAKI', 'PEREMPUAN']
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = bold
    cell.alignment = center
    cell.fill = header_fill
    cell.border = border

start_row = 4
for i, row in enumerate(summary, start=start_row):
    values = [row['no'], row['eselon'], row['jumlah'], row['keterangan'], row['laki_laki'], row['perempuan']]
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=i, column=col, value=value)
        cell.border = border
        cell.alignment = center if col != 4 else left_wrap

jumlah_row = start_row + len(summary)
ws.merge_cells(start_row=jumlah_row, start_column=1, end_row=jumlah_row, end_column=2)
ws.cell(row=jumlah_row, column=1, value='Jumlah Keseluruhan')
ws.cell(row=jumlah_row, column=3, value=total_jumlah)
ws.cell(row=jumlah_row, column=5, value=total_laki)
ws.cell(row=jumlah_row, column=6, value=total_perempuan)
for col in [1, 3, 5, 6]:
    ws.cell(row=jumlah_row, column=col).font = bold
for col in range(1, 7):
    ws.cell(row=jumlah_row, column=col).border = border
    ws.cell(row=jumlah_row, column=col).alignment = center
    ws.cell(row=jumlah_row, column=col).fill = header_fill

for width_col, width in {'A': 8, 'B': 18, 'C': 12, 'D': 72, 'E': 14, 'F': 14}.items():
    ws.column_dimensions[width_col].width = width

for row_num in range(1, jumlah_row + 1):
    ws.row_dimensions[row_num].height = 24

ws2 = wb.create_sheet('Detail Pegawai')
detail_headers = ['Kategori', 'Nama', 'Jenis Kelamin', 'Status Kepegawaian', 'Jabatan Terakhir', 'SKPD']
for col, header in enumerate(detail_headers, start=1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = bold
    cell.alignment = center
    cell.fill = header_fill
    cell.border = border

for row_idx, row in enumerate(detail_rows, start=2):
    values = [row['kategori'], row['nama'], row['jenis_kelamin'], row['status_kepegawaian'], row['jabatan_terakhir'], row['skpd']]
    for col, value in enumerate(values, start=1):
        cell = ws2.cell(row=row_idx, column=col, value=value)
        cell.border = border
        cell.alignment = left_wrap if col in {2, 5, 6} else center

for idx, width in enumerate([18, 32, 16, 18, 70, 60], start=1):
    ws2.column_dimensions[get_column_letter(idx)].width = width

wb.save(OUTPUT)

print(f'Output: {OUTPUT.name}')
print(f'Total ASN dihitung: {total_jumlah}')
print(f'Laki-laki: {total_laki}')
print(f'Perempuan: {total_perempuan}')
for row in summary:
    print(f"{row['eselon']}: jumlah={row['jumlah']}, laki-laki={row['laki_laki']}, perempuan={row['perempuan']}")
