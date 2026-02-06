import json
from openpyxl import Workbook

# ===== 1. Load JSON =====
with open("DUKPNSFULL.json", "r", encoding="utf-8") as f:
    data = json.load(f)

rows = []

for pegawai in data:
    row = {}
    for k, v in pegawai.items():
        if isinstance(v, dict):
            for subk, subv in v.items():
                row[f"{k}_{subk}"] = subv

        elif isinstance(v, list):
            if v:
                latest = v[-1]
                if isinstance(latest, dict):
                    for subk, subv in latest.items():
                        row[f"{k}_{subk}"] = subv
                else:
                    row[k] = latest
            else:
                row[k] = None
        else:
            row[k] = v

    rows.append(row)

# ===== 2. Map JSON keys → Excel column letters =====
column_map = {
    "namaWithGelar": "B",
    "nip": "C",
    "jabatanTerakhir":"L",
    "skpd": "N",
    "riwayatPendidikan_tingkatPendidikan": "S",
    "riwayatPendidikan_jurusan": "T",
    "riwayatPendidikan_namaSekolah": "U",
    "riwayatPendidikan_tanggalIjazah": "V",
    "agama": "AB",
    "jenisKelamin": "AD",
    "golDarah": "AE",
    "noHp": "AF",
    "email": "AG",
    "tempatTglLahir": "AH",
    "alamat": "AI",
    "provinsi": "AJ",
    "kota": "AK",
    "kecamatan": "AL",
    "kelurahan": "AM",
    "rt": "AN",
    "rw": "AO",
}

# ===== 3. Create Excel =====
wb = Workbook()
ws = wb.active
ws.title = "Pegawai"

# headers
for key, col in column_map.items():
    ws[f"{col}1"] = key.upper()

# data rows
for row_idx, row_data in enumerate(rows, start=2):
    for key, col in column_map.items():
        ws[f"{col}{row_idx}"] = row_data.get(key)

# ===== 4. Save =====
wb.save("pegawaiKhusus.xlsx")
print("Excel berhasil dibuat: pegawai.xlsx")
