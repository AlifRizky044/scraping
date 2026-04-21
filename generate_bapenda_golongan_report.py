import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path('/Users/nevv/Documents/scraping')
INPUT = BASE / 'DUKPNSFULLmaret.json'
OUTPUT = BASE / 'Rekap_Jumlah_ASN_BAPENDA_Menurut_Golongan_Maret.xlsx'

EXCLUDED_NAMES = {
    'ANNA SARI HASIBUAN',
}

GOLONGAN_ORDER = [
    'IV/c', 'IV/b', 'IV/a',
    'III/d', 'III/c', 'III/b', 'III/a',
    'II/d', 'II/c', 'II/b', 'II/a',
    'I/d', 'I/c', 'I/b', 'I/a',
]


def normalize_golongan(value: str) -> str:
    value = (value or '').strip().upper()
    if not value or '/' not in value:
        return ''
    major, minor = value.split('/', 1)
    return f'{major}/{minor.lower()}'


with INPUT.open('r', encoding='utf-8') as f:
    data = json.load(f)

rows = [
    row for row in data
    if row.get('statusKepegawaian') in {'PNS', 'CPNS'}
    and (row.get('nama') or '').strip().upper() not in EXCLUDED_NAMES
]

summary = []
detail_rows = []
for idx, gol in enumerate(GOLONGAN_ORDER, start=1):
    members = [row for row in rows if normalize_golongan(row.get('golonganTerakhir', '')) == gol]
    laki = sum(1 for row in members if row.get('jenisKelamin') == 'Laki-laki')
    perempuan = sum(1 for row in members if row.get('jenisKelamin') == 'Perempuan')
    summary.append({
        'no': idx,
        'golongan': gol,
        'jumlah': len(members),
        'laki_laki': laki,
        'perempuan': perempuan,
    })
    for member in members:
        detail_rows.append({
            'golongan': gol,
            'nama': member.get('nama', ''),
            'jenis_kelamin': member.get('jenisKelamin', ''),
            'status_kepegawaian': member.get('statusKepegawaian', ''),
            'jabatan_terakhir': member.get('jabatanTerakhir', ''),
            'skpd': member.get('skpd', ''),
        })

total_jumlah = len(rows)
total_laki = sum(1 for row in rows if row.get('jenisKelamin') == 'Laki-laki')
total_perempuan = sum(1 for row in rows if row.get('jenisKelamin') == 'Perempuan')

wb = Workbook()
ws = wb.active
ws.title = 'Rekap Golongan'

thin = Side(style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill('solid', fgColor='D9EAF7')
title_fill = PatternFill('solid', fgColor='B7DEE8')
bold = Font(bold=True)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

ws.merge_cells('A1:E1')
ws['A1'] = 'Jumlah ASN Badan Pendapatan Daerah Kota Medan Menurut Golongan'
ws['A1'].font = Font(bold=True, size=14)
ws['A1'].alignment = center
ws['A1'].fill = title_fill

headers = ['NO', 'GOLONGAN', 'JUMLAH', 'LAKI-LAKI', 'PEREMPUAN']
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = bold
    cell.alignment = center
    cell.fill = header_fill
    cell.border = border

start_row = 4
for i, row in enumerate(summary, start=start_row):
    values = [row['no'], row['golongan'], row['jumlah'], row['laki_laki'], row['perempuan']]
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=i, column=col, value=value)
        cell.border = border
        cell.alignment = center

jumlah_row = start_row + len(summary)
ws.merge_cells(start_row=jumlah_row, start_column=1, end_row=jumlah_row, end_column=2)
ws.cell(row=jumlah_row, column=1, value='Jumlah Keseluruhan')
ws.cell(row=jumlah_row, column=3, value=total_jumlah)
ws.cell(row=jumlah_row, column=4, value=total_laki)
ws.cell(row=jumlah_row, column=5, value=total_perempuan)
for col in [1, 3, 4, 5]:
    ws.cell(row=jumlah_row, column=col).font = bold
for col in range(1, 6):
    ws.cell(row=jumlah_row, column=col).border = border
    ws.cell(row=jumlah_row, column=col).alignment = center
    ws.cell(row=jumlah_row, column=col).fill = header_fill

for width_col, width in {'A': 8, 'B': 16, 'C': 12, 'D': 14, 'E': 14}.items():
    ws.column_dimensions[width_col].width = width

for row_num in range(1, jumlah_row + 1):
    ws.row_dimensions[row_num].height = 24

ws2 = wb.create_sheet('Detail Pegawai')
detail_headers = ['Golongan', 'Nama', 'Jenis Kelamin', 'Status Kepegawaian', 'Jabatan Terakhir', 'SKPD']
for col, header in enumerate(detail_headers, start=1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = bold
    cell.alignment = center
    cell.fill = header_fill
    cell.border = border

for row_idx, row in enumerate(detail_rows, start=2):
    values = [row['golongan'], row['nama'], row['jenis_kelamin'], row['status_kepegawaian'], row['jabatan_terakhir'], row['skpd']]
    for col, value in enumerate(values, start=1):
        cell = ws2.cell(row=row_idx, column=col, value=value)
        cell.border = border
        cell.alignment = left_wrap if col in {2, 5, 6} else center

for idx, width in enumerate([12, 32, 16, 18, 70, 60], start=1):
    ws2.column_dimensions[get_column_letter(idx)].width = width

wb.save(OUTPUT)

print(f'Output: {OUTPUT.name}')
print(f'Total ASN dihitung: {total_jumlah}')
print(f'Laki-laki: {total_laki}')
print(f'Perempuan: {total_perempuan}')
for row in summary:
    print(f"{row['golongan']}: jumlah={row['jumlah']}, laki-laki={row['laki_laki']}, perempuan={row['perempuan']}")
